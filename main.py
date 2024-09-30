import os
import requests
import streamlit as st
import docx2txt
import PyPDF2
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import langdetect


# בדיקה אם אנחנו בסביבת Streamlit Cloud
STREAMLIT_DEPLOYMENT = os.getenv('STREAMLIT_DEPLOYMENT', 'false').lower() == 'true'

# ניסיון לייבא ChromeDriverManager רק אם לא בסביבת Streamlit Cloud
if not STREAMLIT_DEPLOYMENT:
    try:
        from webdriver_manager.chrome import ChromeDriverManager
    except ImportError:
        ChromeDriverManager = None
else:
    ChromeDriverManager = None

def create_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    
   # בדיקה אם אנחנו בסביבת ייצור או Streamlit Cloud
    if os.getenv('ENVIRONMENT') == 'production' or STREAMLIT_DEPLOYMENT:
        service = Service('/usr/bin/chromedriver')
    else:
        # בסביבת פיתוח, השתמש ב-ChromeDriverManager אם הוא זמין
        if ChromeDriverManager:
            service = Service(ChromeDriverManager().install())
        else:
            # אם ChromeDriverManager לא זמין, נסה להשתמש בדרייבר מקומי
            service = Service('chromedriver')
    
    return webdriver.Chrome(service=service, options=chrome_options)

# שאר הקוד נשאר ללא שינוי
import io
import pandas as pd
from dotenv import load_dotenv
import urllib.parse
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import base64

load_dotenv()

def load_resume(file):
    file_type = file.type
    if file_type == "application/pdf":
        return read_pdf(file)
    elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        return read_docx(file)
    else:
        st.error("פורמט קובץ לא נתמך. אנא העלה קובץ PDF או Word.")
        return None

def read_pdf(file):
    pdf_reader = PyPDF2.PdfReader(io.BytesIO(file.getvalue()))
    text = ""
    for page in pdf_reader.pages:
        text += page.extract_text()
    return text

def read_docx(file):
    return docx2txt.process(io.BytesIO(file.getvalue()))

# טעינת הסוכנים מקובץ JSON
with open('agents.json', 'r', encoding='utf-8') as file:
    agents = json.load(file)

def request_url(agent, prompt, page=1):
    driver = create_driver()
    base_url = agent['url']
    start_text = agent['start_text']
    end_text = agent['end_text']
    
    encoded_prompt = urllib.parse.quote(prompt)
    if '{page}' in base_url:
        url = base_url.format(prompt=encoded_prompt, page=page)
    else:
        url = base_url.format(prompt=encoded_prompt)
    
    try:
        driver.get(url)
        content = driver.page_source
        
        start_index = content.find(start_text)
        end_index = content.find(end_text, start_index)

        if start_index != -1 and end_index != -1:
            extracted_content = content[start_index:end_index + len(end_text)]
            return extracted_content
        else:
            st.warning(f"לא נמצא תוכן מתאים עבור {agent['name']}")
            return None

    except Exception as e:
        st.error(f"אירעה שגיאה בעת שליפת נתונים מ-{agent['name']}: {str(e)}")
        return None
    finally:
        driver.quit()

def extract_jobs_drushim(html_content, agent_name):
    soup = BeautifulSoup(html_content, 'html.parser')
    jobs = []
    
    for job_div in soup.find_all('div', class_='job-item-main'):
        job = {'source': agent_name}  # Add the agent name to each job
        
        # תפקיד
        title = job_div.find('h3', class_='display-28')
        job['title'] = title.text.strip() if title else ''
        
        # חברה
        company = job_div.find('p', class_='display-22')
        job['company'] = company.text.strip() if company else ''
        
        # מיקום, שנות ניסיון, היקף משרה, פורסם לפני
        details = job_div.find_all('span', class_='display-18')
        for detail in details:
            text = detail.text.strip()
            if 'שנים' in text:
                job['experience'] = text
            elif any(word in text for word in ['משרה מלאה', 'משרה חלקית']):
                job['job_type'] = text
            elif 'לפני' in text:
                job['posted'] = text
            else:
                job['location'] = text
        
        jobs.append(job)
    
    return jobs
def extract_jobs_jobmaster(html_content, agent_name):
    soup = BeautifulSoup(html_content, 'html.parser')
    jobs = []
    
    for job_div in soup.find_all('article', class_='JobItem'):
        job = {'source': agent_name}  # Add the agent name to each job
        
        # תפקיד (Title)
        title = job_div.find('a', class_='CardHeader')
        job['title'] = title.text.strip() if title else ''
        
        # חברה (Company)
        company = job_div.find('a', class_='CompanyNameLink') or job_div.find('span', class_='ByTitle')
        job['company'] = company.text.strip() if company else ''
        
        # מיקום (Location)
        location = job_div.find('li', class_='jobLocation')
        job['location'] = location.text.strip() if location else ''
        
        # שנות ניסיון (Experience)
        # JobMaster doesn't seem to have a specific field for experience, so we'll leave it empty
        job['experience'] = ''
        
        # היקף משרה (Job Type)
        job_type = job_div.find('li', class_='jobType')
        job['job_type'] = job_type.text.strip() if job_type else ''
        
        # פורסם לפני (Posted)
        posted = job_div.find('span', class_='Gray')
        job['posted'] = posted.text.strip() if posted else ''
        
        # תיאור קצר (Short Description)
        description = job_div.find('div', class_='jobShortDescription')
        job['description'] = description.text.strip() if description else ''
        
        jobs.append(job)
    
    return jobs
def extract_jobs_avodata(html_content, agent_name):
    soup = BeautifulSoup(html_content, 'html.parser')
    jobs = []
    
    for job_div in soup.find_all('div', class_='result-TaasukaCatalog'):
        job = {'source': agent_name}
        
        # תפקיד (Title)
        title = job_div.find('div', class_='title')
        job['title'] = title.text.strip() if title else ''
        
        # תיאור (Description)
        description = job_div.find('div', class_='sub-title')
        job['description'] = description.text.strip() if description else ''
        
        # מידע נוסף
        info_list = job_div.find('ul')
        if info_list:
            for li in info_list.find_all('li'):
                if 'belongsToScope' in li.get('class', []):
                    job['field'] = li.text.split(': ')[1] if ': ' in li.text else ''
                elif 'salary' in li.get('class', []):
                    job['salary'] = li.text.split(': ')[1] if ': ' in li.text else ''
        
        # לינק למשרה
        link = job_div.find('a', class_='result-TaasukaCatalog')
        job['link'] = link['href'] if link and 'href' in link.attrs else ''
        
        # שדות שאינם זמינים ב-Avodata
        job['company'] = ''
        job['location'] = ''
        job['experience'] = ''
        job['job_type'] = ''
        job['posted'] = ''
        
        jobs.append(job)
    
    return jobs

def create_excel_from_json(jobs):
    wb = Workbook()
    ws = wb.active
    ws.title = "רשימת משרות"

    headers = ['מקור', 'תפקיד', 'חברה', 'מיקום', 'ניסיון', 'סוג משרה', 'פורסם']

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, job in enumerate(jobs, start=2):
        ws.cell(row=row, column=1, value=job.get('source', ''))
        ws.cell(row=row, column=2, value=job.get('title', ''))
        ws.cell(row=row, column=3, value=job.get('company', ''))
        ws.cell(row=row, column=4, value=job.get('location', ''))
        ws.cell(row=row, column=5, value=job.get('experience', ''))
        ws.cell(row=row, column=6, value=job.get('job_type', ''))
        ws.cell(row=row, column=7, value=job.get('posted', ''))

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    return wb

def get_table_download_link(df):
    """Generates a link allowing the data in a given panda dataframe to be downloaded"""
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    writer.save()
    processed_data = output.getvalue()
    b64 = base64.b64encode(processed_data).decode()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="job_listings.xlsx">Download Excel file</a>'

def detect_language(text):
    try:
        return langdetect.detect(text)
    except:
        return 'en'  # ברירת מחדל לאנגלית אם הזיהוי נכשל

def analyze_jobs_with_groq(resume, jobs, language):
    api_key = os.getenv('GROQ_API_KEY')
    if not api_key:
        st.error("GROQ API key is missing. Please set it in your .env file.")
        return []

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    prompt = f"""
    Given the following resume and list of job postings, analyze and rank the top 5 jobs that best match the candidate's skills and experience. For each job, provide a brief explanation of why it's a good match.

    The resume is in {language} language. Please provide your response in {language}.

    Resume:
    {resume}

    Job Postings:
    {json.dumps(jobs, ensure_ascii=False)}

    Please provide the results in the following format, using {language}:
    1. Job Title (Company Name)
       Explanation: [Brief explanation of why this job is a good match]

    2. Job Title (Company Name)
       Explanation: [Brief explanation of why this job is a good match]

    ... and so on for the top 5 matches.
    """

    data = {
        "model": "mixtral-8x7b-32768",
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.5,
        "max_tokens": 1000
    }

    response = requests.post("https://api.groq.com/openai/v1/chat/completions", headers=headers, json=data)
    
    if response.status_code == 200:
        return response.json()['choices'][0]['message']['content']
    else:
        st.error(f"Error from Groq API: {response.text}")
        return []
    
def main():
    st.set_page_config(layout="wide", page_title="סוכן משרות מקצועי", page_icon="🔍")

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Heebo:wght@400;700&display=swap');
    
    body {
        direction: rtl;
        text-align: right;
        font-family: 'Heebo', sans-serif;
    }
    .stButton>button {
        width: 100%;
    }
    .stSelectbox>div>div>select {
        direction: rtl;
    }
    </style>
    """, unsafe_allow_html=True)

    st.title("סוכן משרות מקצועי")

    selected_agents = st.multiselect("בחר סוכנים", [agent['name'] for agent in agents], default=[agent['name'] for agent in agents])
    prompt = st.text_input("הכנס מילות חיפוש", value="מנהל שיווק ומכירות")
    
    uploaded_file = st.file_uploader("העלה קורות חיים (PDF או Word)", type=["pdf", "docx"])
    
    # הוספת אפשרות בחירת שפה
    language_options = {
        "עברית": "he",
        "אנגלית": "en",
        "רוסית": "ru",
        "ערבית": "ar",
        "צרפתית": "fr"
    }
    selected_language = st.selectbox("בחר את השפה שבה אתה רוצה לקבל תשובה", options=list(language_options.keys()))

    if st.button("חפש משרות שמתאימות לקורות חיים שלי"):
        if not selected_agents:
            st.warning("אנא בחר לפחות סוכן אחד")
            return

        progress_bar = st.progress(0)
        status_text = st.empty()

        all_jobs = []
        for i, agent_name in enumerate(selected_agents):
            agent = next((a for a in agents if a['name'] == agent_name), None)
            if agent:
                status_text.text(f"מחפש משרות ב-{agent_name}...")
                html_content = request_url(agent, prompt=prompt, page=1)
                
                if html_content:
                    extract_function = globals()[f"extract_jobs_{agent['name']}"]
                    extracted_jobs = extract_function(html_content, agent['name'])
                    all_jobs.extend(extracted_jobs)
                
                progress_bar.progress((i + 1) / len(selected_agents))

        if not all_jobs:
            st.warning("לא נמצאו משרות מתאימות")
            return

        status_text.text("יוצר קובץ Excel...")
        wb = create_excel_from_json(all_jobs)
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        b64 = base64.b64encode(excel_file.read()).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="רשימת_משרות.xlsx">לחץ כאן להורדת קובץ Excel</a>'
        st.markdown(href, unsafe_allow_html=True)

        status_text.text(f"נמצאו {len(all_jobs)} משרות בסך הכל.")
        progress_bar.progress(100)

        try:
            df = pd.DataFrame(all_jobs)
            st.write("תוצאות החיפוש:")
            st.dataframe(df)

            if uploaded_file is not None:
                resume_content = load_resume(uploaded_file)
                if resume_content:
                    status_text.text("מנתח התאמה לקורות החיים...")
                    # שימוש בשפה שנבחרה
                    matching_results = analyze_jobs_with_groq(resume_content, all_jobs, language_options[selected_language])
                    st.subheader("המשרות המתאימות ביותר לקורות החיים שלך:")
                    st.write(matching_results)
            
        except Exception as e:
            st.error(f"אירעה שגיאה ביצירת טבלת הנתונים: {str(e)}")
            st.write("Debug: תוכן all_jobs:")
            st.write(all_jobs)

if __name__ == "__main__":
    main()